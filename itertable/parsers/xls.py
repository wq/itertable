import datetime
import math
from .base import TableParser


class WorkbookParser(TableParser):
    workbook = None
    worksheet = None
    sheet_name = 0
    start_row = None
    column_count = None
    no_pickle_parser = ["workbook", "worksheet"]
    binary = True

    date_format = "yyyy-mm-dd"
    time_format = "hh:mm:ss"
    datetime_format = "yyyy-mm-dd hh:mm:ss"

    def parse(self):
        if not self.workbook:
            self.parse_workbook()

        if self.sheet_name is None:
            SpreadsheetIter = type(self)
            self.data = [
                {
                    "name": name,
                    "data": SpreadsheetIter(
                        loaded=True,
                        workbook=self.workbook,
                        sheet_name=name,
                    ),
                }
                for name in self.sheet_names
            ]
            return

        sheet_name = self.sheet_name
        if isinstance(self.sheet_name, int):
            sheet_name = self.sheet_names[sheet_name]

        self.parse_worksheet(sheet_name)

        if self.header_row is None:
            if self.start_row is not None:
                self.header_row = self.start_row - 1
            else:
                self.column_count = 0

                def checkval(cell):
                    if cell.value is not None and cell.value != "":
                        return True
                    return False

                search_rows = min(len(self.worksheet) - 1, self.max_header_row)
                for row in range(search_rows, -1, -1):
                    count = len(list(filter(checkval, self.worksheet[row])))
                    if count >= self.column_count:
                        self.column_count = count
                        self.header_row = row

        if self.header_row is None:
            return

        if self.start_row is None:
            self.start_row = self.header_row + 1

        if self.field_names is None:
            rows = self.worksheet[self.header_row : self.start_row]
            self.field_names = [
                str(c.value) or "c%s" % i for i, c in enumerate(rows[0])
            ]
            for row in rows[1:]:
                for i, c in enumerate(row):
                    self.field_names[i] += "\n" + str(c.value)

            seen_fields = set()
            for i, field in enumerate(self.field_names):
                if field in seen_fields:
                    field += str(i)
                    self.field_names[i] = field
                seen_fields.add(field)

        self.data = list(map(self.parse_row, self.worksheet[self.start_row :]))

        self.extra_data = {}
        if self.header_row > 0:
            for r in range(0, self.header_row):
                for c, cell in enumerate(self.worksheet[r]):
                    val = self.get_value(cell)
                    if val is not None and val != "":
                        self.extra_data.setdefault(r, {})
                        self.extra_data[r][c] = val

    def parse_workbook(self):
        raise NotImplementedError

    @property
    def sheet_names(self):
        raise NotImplementedError

    def get_sheet_by_name(self, name):
        raise NotImplementedError

    def parse_worksheet(self, name):
        raise NotImplementedError

    def parse_row(self, row):
        return {
            name: self.get_value(row[i])
            for i, name in enumerate(self.get_field_names())
            if i < len(row)
        }

    def get_value(self, cell):
        raise NotImplementedError

    def dump(self, file=None):
        if file is None:
            file = self.file
        write, close = self.open_worksheet(file)
        for i, field in enumerate(self.field_names):
            write(0, i, field)
        for r, row in enumerate(self.data):
            for c, field in enumerate(self.field_names):
                write(r + 1, c, row[field])
        close()

    def calc_width(self, val):
        val = str(val) if val is not None else ""
        size = 0
        for c in val:
            if c in ".,;:'\"iIlt1":
                size += 0.5
            elif c in "MW":
                size += 1.3
            elif c.isupper():
                size += 1.2
            elif c.islower():
                size += 1
            else:
                size += 1.1
        return size * 1.4


class OldExcelParser(WorkbookParser):
    def parse_workbook(self):
        try:
            import xlrd
        except ImportError:
            raise Exception("xlrd is required to load .xls files")
        self.workbook = xlrd.open_workbook(file_contents=self.file.read())

    @property
    def sheet_names(self):
        return self.workbook.sheet_names()

    def get_sheet_by_name(self, name):
        return self.workbook.sheet_by_name(name)

    def parse_worksheet(self, name):
        worksheet = self.get_sheet_by_name(name)
        self.worksheet = [worksheet.row(i) for i in range(worksheet.nrows)]

    def get_value(self, cell):
        import xlrd

        if cell.ctype == xlrd.XL_CELL_DATE:
            time, date = math.modf(cell.value)
            tpl = xlrd.xldate_as_tuple(cell.value, self.workbook.datemode)
            if date and time:
                return datetime.datetime(*tpl)
            elif date:
                return datetime.date(*tpl[0:3])
            else:
                return datetime.time(*tpl[3:6])
        return cell.value

    def calc_width(self, val):
        val = str(val) if val is not None else ""
        size = 0
        for c in val:
            if c in ".,;:'\"iIlt1":
                size += 0.5
            elif c in "MW":
                size += 1.3
            elif c.isupper():
                size += 1.2
            elif c.islower():
                size += 1
            else:
                size += 1.1
        return size

    def open_worksheet(self, file):
        import xlwt

        workbook = xlwt.Workbook()
        worksheet = workbook.add_sheet("Sheet 1")

        formats = {
            datetime.date: xlwt.Style.easyxf(
                num_format_str=self.date_format,
            ),
            datetime.time: xlwt.Style.easyxf(
                num_format_str=self.time_format,
            ),
            datetime.datetime: xlwt.Style.easyxf(
                num_format_str=self.datetime_format,
            ),
            "header": xlwt.Style.easyxf(
                "font: bold on; borders: bottom thick;"
            ),
        }

        widths = {}

        def write(r, c, val):
            widths.setdefault(c, 0)
            widths[c] = max(widths[c], self.calc_width(val))
            fmt = formats.get(type(val))
            if not fmt and r == 0:
                fmt = formats["header"]
            if fmt:
                worksheet.write(r, c, val, fmt)
            else:
                worksheet.write(r, c, val)

        def close():
            for c, width in widths.items():
                worksheet.col(c).set_width(int(width * 256))
            workbook.save(file)

        return write, close


class ExcelParser(WorkbookParser):
    def parse_workbook(self):
        import openpyxl

        self.workbook = openpyxl.open(self.file, data_only=True)

    @property
    def sheet_names(self):
        return self.workbook.sheetnames

    def get_sheet_by_name(self, name):
        return self.workbook[name]

    def parse_worksheet(self, name):
        worksheet = self.get_sheet_by_name(name)
        self.worksheet = [row for row in worksheet.rows]

    def get_value(self, cell):
        value = cell.internal_value
        if isinstance(value, datetime.datetime):
            if value.time() == datetime.time(0, 0):
                return value.date()
        return value

    def open_worksheet(self, file):
        from openpyxl import Workbook, styles, utils

        workbook = Workbook()
        worksheet = workbook.active

        formats = {
            datetime.date: styles.NamedStyle(
                name="date",
                number_format=self.date_format,
            ),
            datetime.time: styles.NamedStyle(
                name="time",
                number_format=self.time_format,
            ),
            datetime.datetime: styles.NamedStyle(
                name="datetime",
                number_format=self.datetime_format,
            ),
            "header": styles.NamedStyle(
                name="header",
                font=styles.Font(bold=True),
                border=styles.Border(bottom=styles.Side(style="thick")),
            ),
        }
        widths = {}

        def write(r, c, val):
            widths.setdefault(c, 0)
            widths[c] = max(widths[c], self.calc_width(val))
            cell = worksheet.cell(r + 1, c + 1, val)

            fmt = formats.get(type(val))
            if fmt:
                cell.style = fmt
            elif r == 0:
                cell.style = formats["header"]

        def close():
            for c, width in widths.items():
                col = utils.get_column_letter(c + 1)
                worksheet.column_dimensions[col].width = width
            workbook.save(self.filename)

        return write, close

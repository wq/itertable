from wq.io.base import IO
from wq.io.file import BinaryFileIO

class _XlBaseIO(IO):
    @property
    def sheet_names(self):
        raise NotImplementedError

    @property
    def get_sheet_by_name(self, name):
        raise NotImplementedError

    def __iter__(self):
        for name in self.sheet_names:
            yield name

    def __getitem__(self, key):
        data = self.get_sheet_by_name(key)
        if data is not None:
            return SheetIO(book=self, data=data)
        else:
            raise KeyError

# Deal with pre-2003 format files (.xls)
class _OldXlIO(_XlBaseIO):
    def load(self):
        import xlrd
        self.data = xlrd.open_workbook(file_contents=self.file.read())

    @property
    def sheet_names(self):
        return self.data.sheet_names()

    def get_sheet_by_name(self, name):
        return self.data.sheet_by_name(name)

# Deal with 2003+ format files (.xlsx)
class _NewXlIO(_XlBaseIO):
    def load(self):
        import openpyxl
        self.data = openpyxl.reader.excel.load_workbook(self.file, use_iterators=True)

    @property
    def sheet_names(self):
        return self.data.get_sheet_names()

    def get_sheet_by_name(self, name):
        return self.data.get_sheet_by_name(name)

class XlIO(_NewXlIO):
    pass

class SheetIO(IO):
    book = None

    def open(self):
        pass

    def load(self):
        pass
    
    @property
    def field_names(self):
        for row in self.data.iter_rows():
            return [c.internal_value or c.column for c in row]

    def __iter__(self):
        for row in self.data.iter_rows():
            yield self.totuple(row)

    def __getitem__(self, key):
        raise NotImplementedError

    def totuple(self, row):
        data = [c.internal_value for c in row]
        return self.item_class._make(data)

class XlFileIO(XlIO, BinaryFileIO):
    pass
